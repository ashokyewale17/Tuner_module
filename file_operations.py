import openpyxl
import xlsxwriter
from datetime import datetime
from tkinter import filedialog

class FileOperations:
    def __init__(self, gui):
        self.gui = gui

    def open_file(self):
        filename = filedialog.askopenfilename(title="Open a File", filetype=(("xlxs files", ".*xlsx"), ("All Files", "*.")))
        if filename:
            try:
                filename = r"{}".format(filename)
                df = openpyxl.load_workbook(filename)
                sheet = df.active
                max_row = sheet.max_row - 2
                rE = 0
                le = 0
                for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=8):
                    row_data = []
                    for cell in row:
                        if cell.column_letter == "B" or cell.column_letter == "D" or cell.column_letter == "F" or cell.column_letter == "H":
                            row_data.append(cell.value)

                    if le < len(self.gui.entry_lists1):
                        self.gui.data_handler.AlgoToSetValue(self.gui.entry_lists1[le], row_data)
                    else:
                        print(f"Warning: entry_lists1 index {le} is out of range.")
                    rE += 1
                    le += 1

                self.gui.dialog.configure(text="Upload successful", font=('Helvetica', 17, 'bold'), fg_color="#ADD8E6")

            except ValueError:
                self.gui.dialog.configure(text="File could not be opened")
            except FileNotFoundError:
                self.gui.dialog.configure(text="File Not Found")

    def write_excelbutton(self):
        now = datetime.now()
        default_filename = 'MCU_Parameters' + now.strftime("D%d_%m_%Y_T%H-%M") + '.xlsx'
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save the Excel file", initialfile=default_filename)

        if save_path:
            workbook = xlsxwriter.Workbook(save_path)
            worksheet = workbook.add_worksheet()

            column_letters = ["A", "B", "C", "D", "E", "F", "G", "H"]
            col, worksheet = self.common_write('full_config.json', column_letters, 1, worksheet, self.gui.entry_lists1)

            last_row_index = col + 1
            worksheet.write(f"A{last_row_index + 1}", "Date")
            worksheet.write(f"B{last_row_index + 1}", now.strftime("%d/%m/%Y %H:%M:%S"))
            worksheet.write(f"A{last_row_index + 2}", "PicID")
            worksheet.write(f'B{last_row_index + 2}', self.gui.label2.get())
            worksheet.write(f'C{last_row_index + 2}', self.gui.label3.get())
            worksheet.write(f'A{last_row_index + 3}', "Firmware Version")
            worksheet.write(f'B{last_row_index + 3}', self.gui.firmware2.get())
            worksheet.write(f'C{last_row_index + 3}', self.gui.firmware3.get())

            workbook.close()

    def common_write(self, data, list, col, worksheet, entry_list):
        filename = self.gui.data_handler.get_json_file_path(data)
        with open(filename, 'r') as f:
            data = json.load(f)
            value = data['instructions']
            le = 0
            for i in range(len(value)):
                ind = value[i]['fields']
                row = 0
                for j in range(len(ind)):
                    worksheet.write(list[row] + str(col), ind[j]['field1_name'])
                    row += 2
                row = 1
                for j in entry_list[le]:
                    if isinstance(j, str):
                        continue
                    else:
                        worksheet.write(list[row] + str(col), float(j.get()))
                        row += 2
                le += 1
                col += 1
        return col, worksheet