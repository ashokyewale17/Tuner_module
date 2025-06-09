import os
import sys
import re
import json
import serial
import threading
import time
import struct
import random
import pyautogui
import xlsxwriter
import openpyxl
import binascii
import webbrowser
import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
from tkinter import filedialog
import serial.tools.list_ports
from datetime import datetime
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinter import simpledialog, messagebox


# get the path to the script directory
script_dir = sys.path[0]

class MainGui:
    def __init__(self, window):
        self.root = window
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        data = 'full_config.json'
        filename = self.get_json_file_path(data)
        self.uartState = False
        self.is_running = False
        self.is_action_in_progress = False
        self.start_time = time.time()
        self.data = None
        self.entry_lists1 = []
        self.flag = 0
        self.flag2 = 0
        self.open = 0
        self.ser = None
        self.le = 0
        self.ser = serial.Serial()
        self.selected_port = tk.StringVar()
        self.button_zero = None

        def resource_path(relative_path):
            """ Get the absolute path to the resource, works for dev and for PyInstaller """
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            return os.path.join(base_path, relative_path)

        icon_path = resource_path('aptech.ico')  # Path to your .ico file
        root.iconbitmap(icon_path)

        # Set modern theme and appearance
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        self.setup_gui()
        
        # Serial object
        self.ser = serial.Serial()
        self.root.mainloop()

    def setup_gui(self):
        # Main window configuration - Make it responsive
        self.root.title("Tuner Software V6.0")
        self.root.resizable(True, True)
        self.root.configure(fg_color="white")
        
        # Set window size and position based on screen resolution
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Responsive window sizing
        if screen_width >= 3840:  # 4K and above
            window_width = int(screen_width * 0.7)
            window_height = int(screen_height * 0.8)
        elif screen_width >= 1920:  # Full HD
            window_width = int(screen_width * 0.85)
            window_height = int(screen_height * 0.85)
        elif screen_width >= 1366:  # 1366x768 
            window_width = int(screen_width * 0.95)  
            window_height = int(screen_height * 0.90)  
        else:  # 1024x768 and similar
            window_width = int(screen_width * 0.95)
            window_height = int(screen_width * 0.9)
        
        # Minimum size constraints - adjusted for 1366x768
        if screen_width >= 1366:
            window_width = max(1300, window_width)
            window_height = max(690, window_height)  # Reduced for 1366x768
        else:
            window_width = max(1200, window_width)
            window_height = max(800, window_height)

        
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        # Set minimum size based on screen resolution
        if screen_width >= 1366:
            self.root.minsize(1300, 690)
        else:
            self.root.minsize(1000, 650)
        
        # Configure root grid for responsiveness
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Create main container with grid layout
        self.main_container = ctk.CTkFrame(self.root, fg_color="white")
        self.main_container.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        # Configure main container grid
        self.main_container.grid_rowconfigure(0, weight=0)  # Header - fixed
        self.main_container.grid_rowconfigure(1, weight=0)  # Control panel - fixed
        self.main_container.grid_rowconfigure(2, weight=1)  # Config area - expandable
        self.main_container.grid_rowconfigure(3, weight=0)  # Action buttons - fixed
        self.main_container.grid_columnconfigure(0, weight=1)

        # Create header with responsive design
        self.create_responsive_header()
        
        # Create control panel with responsive design
        self.create_responsive_control_panel()
        
        # Create responsive configuration area
        self.create_responsive_config_area()
        
        # Create responsive action buttons
        self.create_responsive_action_buttons()

    def create_responsive_header(self):
        """Create responsive header section"""
        header_frame = ctk.CTkFrame(self.main_container, fg_color="white")
        header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=(0, 10))
        header_frame.grid_columnconfigure(1, weight=1)  # Make title area expandable
        
        # Logo section
        logo_frame = ctk.CTkFrame(header_frame, fg_color="white")
        logo_frame.grid(row=0, column=0, sticky="w", padx=10)
        
        img_path = self.get_json_file_path('APT_LOGO.png')
        try:
            image = Image.open(img_path)
            # Responsive logo sizing
            screen_width = self.root.winfo_screenwidth()
            if screen_width >= 3840:  # 4K
                width, height = 600, 55
            elif screen_width >= 1920:  # Full HD
                width, height = 500, 45
            elif screen_width >= 1366:  # 1366x768
                width, height = 450, 40
            else:  # Smaller screens
                width, height = 400, 35
                
            resized_image = image.resize((width, height), Image.LANCZOS)
            img = ctk.CTkImage(dark_image=resized_image, size=(width, height))
            logo_label = ctk.CTkLabel(logo_frame, image=img, text="")
            logo_label.pack()
        except Exception as e:
            print(f"Error loading logo: {e}")
        
        # Title section - expandable
        title_frame = ctk.CTkFrame(header_frame, fg_color="white")
        title_frame.grid(row=0, column=1, sticky="ew", padx=15)
        
        # Responsive font sizing
        screen_width = self.root.winfo_screenwidth()
        if screen_width >= 3840:
            title_font_size = 18
            version_font_size = 16
        elif screen_width >= 1920:
            title_font_size = 14
            version_font_size = 13
        elif screen_width >= 1366:  # 1366x768 optimization
            title_font_size = 13
            version_font_size = 12
        else:
            title_font_size = 12
            version_font_size = 11
        
        title_label = ctk.CTkLabel(title_frame, text="⚡TUNER APT SOFTWARE", 
                                 font=('Helvetica', title_font_size, 'bold'), text_color="#2a5885")
        title_label.pack(anchor='w')
        
        version_label = ctk.CTkLabel(title_frame, text="  Version 6.0", 
                                   font=('Helvetica', version_font_size, 'bold'), text_color="black")
        version_label.pack(anchor='w')
        
        # Help button
        help_btn = ctk.CTkButton(header_frame, text="?", width=30, height=30,
                               font=('Helvetica', 14, 'bold'), fg_color="#2a5885", 
                               hover_color="#3b7cb1", command=self.show_help)
        help_btn.grid(row=0, column=2, sticky="e", padx=10)

    def create_responsive_control_panel(self):
        """Create responsive control panel"""
        control_frame = ctk.CTkFrame(self.main_container, border_width=1, 
                                   border_color="#dddddd", corner_radius=8, fg_color="#f8f9fa")
        control_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=1)
        
        # Configure control frame grid for responsive sections
        control_frame.grid_columnconfigure(0, weight=1)  # Serial port
        control_frame.grid_columnconfigure(1, weight=1)  # PIC ID
        control_frame.grid_columnconfigure(2, weight=2)  # Status (wider)
        control_frame.grid_columnconfigure(3, weight=1)  # Firmware
        control_frame.grid_columnconfigure(4, weight=1)     # Color code
        
        # Serial Port Section
        self.create_serial_port_section(control_frame, 0)
        
        # PIC ID Section
        self.create_pic_id_section(control_frame, 1)
        
        # Status Section
        self.create_status_section(control_frame, 2)
        
        # Firmware Section
        self.create_firmware_section(control_frame, 3)

        # Color Code Section
        self.create_color_code_section(control_frame, 4)

    def create_serial_port_section(self, parent, column):
        """Create responsive serial port section"""
        port_frame = ctk.CTkFrame(parent, fg_color="transparent", border_color='black', border_width=2, height=115)
        port_frame.grid(row=0, column=column, sticky="nsew", padx=5, pady=5)
        port_frame.grid_columnconfigure(0, weight=1)
        port_frame.grid_propagate(False)
        
        # Responsive font size
        font_size = 14 if self.root.winfo_screenwidth() >= 1920 else (13 if self.root.winfo_screenwidth() >= 1366 else 12)
        
        ctk.CTkLabel(port_frame, text="SERIAL PORT", font=('Helvetica', font_size),
                   text_color="#333333").grid(row=0, column=0, padx=3, pady=5, sticky="w")
        
        # COM port selection
        serial_ports = self.detect_serial_ports()
        self.selected_port = tk.StringVar(value="Select Port")
        
        port_combo = ctk.CTkComboBox(port_frame, variable=self.selected_port, values=serial_ports, 
                                   dropdown_fg_color="#f8f9fa", dropdown_text_color="#333333", 
                                   button_color="#2a5885", border_color="#cccccc")
        port_combo.grid(row=1, column=0, pady=5, padx=10, sticky="ew")
        self.selected_port.trace_add("write", self.on_port_selected)
        
        # Connection buttons
        btn_frame = ctk.CTkFrame(port_frame, fg_color="white")
        btn_frame.grid(row=2, column=0, padx=3, pady=3, sticky="ew")
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)
        
        self.buttonSS = ctk.CTkButton(btn_frame, text="CONNECT", font=('Helvetica', 11, 'bold'),
                                    fg_color="#28a745", hover_color="#218838", state="disabled", width=90,
                                    height=30, command=self.processButtonSS)
        self.buttonSS.grid(row=0, column=0, padx=5, sticky="ew")
        
        self.buttonStop = ctk.CTkButton(btn_frame, text="DISCONNECT", font=('Helvetica', 11, 'bold'),
                                      fg_color="#dc3545", hover_color="#c82333", state="disabled", width=90,
                                      height=30, command=self.stopButtonSS)
        self.buttonStop.grid(row=0, column=1, padx=5, sticky="ew")

    def create_pic_id_section(self, parent, column):
        """Create responsive PIC ID section"""
        pic_frame = ctk.CTkFrame(parent, fg_color="transparent", border_color='black', border_width=2)
        pic_frame.grid(row=0, column=column, sticky="nsew", padx=10, pady=5)
        pic_frame.grid_columnconfigure(0, weight=1)
        pic_frame.grid_columnconfigure(1, weight=1)
        
        font_size = 15 if self.root.winfo_screenwidth() >= 1920 else (14 if self.root.winfo_screenwidth() >= 1366 else 12)
        
        ctk.CTkLabel(pic_frame, text="PIC ID", font=('Helvetica', font_size, 'bold'), 
                   text_color="#333333").grid(row=0, column=0, columnspan=2, pady=5)
        
        ctk.CTkLabel(pic_frame, text="High", font=('Helvetica', 11, 'bold')).grid(row=1, column=0, padx=5)
        self.label2 = ctk.CTkEntry(pic_frame, height=35, width=100, border_color="black", font=('Helvetica', 11))
        self.label2.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
        
        ctk.CTkLabel(pic_frame, text="Low", font=('Helvetica', 11, 'bold')).grid(row=1, column=1, padx=5)
        self.label3 = ctk.CTkEntry(pic_frame, height=35, width=100, border_color="black", font=('Helvetica', 11))
        self.label3.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

    def create_status_section(self, parent, column):
        """Create responsive status section"""
        status_frame = ctk.CTkFrame(parent, fg_color="transparent", border_color='black', border_width=2)
        status_frame.grid(row=0, column=column, sticky="nsew", padx=10, pady=5)
        status_frame.grid_columnconfigure(0, weight=1)
        status_frame.grid_rowconfigure(1, weight=1)
        
        font_size = 15 if self.root.winfo_screenwidth() >= 1920 else 13
        
        ctk.CTkLabel(status_frame, text="STATUS", font=('Helvetica', font_size, 'bold'), 
                   text_color="#333333").grid(row=0, column=0, pady=10, padx=5, sticky="ew")
        
        self.dialog = ctk.CTkLabel(status_frame, text="", font=('Helvetica', 11), 
                                 anchor="center", justify="center", height=20, width=200, wraplength=300)
        self.dialog.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

    def create_firmware_section(self, parent, column):
        """Create responsive firmware section"""
        firmware_frame = ctk.CTkFrame(parent, fg_color="transparent", border_color='black', border_width=2)
        firmware_frame.grid(row=0, column=column, sticky="nsew", padx=10, pady=5)
        firmware_frame.grid_columnconfigure(0, weight=1)
        firmware_frame.grid_columnconfigure(1, weight=1)
        
        font_size = 15 if self.root.winfo_screenwidth() >= 1920 else (14 if self.root.winfo_screenwidth() >= 1366 else 12)
        
        ctk.CTkLabel(firmware_frame, text="FIRMWARE VERSION", font=('Helvetica', font_size, 'bold'), 
                   text_color="#333333").grid(row=0, column=0, columnspan=2, pady=5)
        
        ctk.CTkLabel(firmware_frame, text="High", font=('Helvetica', 11, 'bold')).grid(row=1, column=0, padx=5)
        self.firmware2 = ctk.CTkEntry(firmware_frame, height=35, width=100, border_color="black", font=('Helvetica', 11))
        self.firmware2.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
        
        ctk.CTkLabel(firmware_frame, text="Low", font=('Helvetica', 11, 'bold')).grid(row=1, column=1, padx=5)
        self.firmware3 = ctk.CTkEntry(firmware_frame, height=35, width=100, border_color="black", font=('Helvetica', 11))
        self.firmware3.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

    def create_color_code_section(self, parent, column):
        color_frame = ctk.CTkFrame(parent, fg_color="transparent", border_color='black', border_width=2)
        color_frame.grid(row=0, column=column, sticky="s", padx=5, pady=5)
        color_frame.grid_columnconfigure(0, weight=1)
        color_frame.grid_columnconfigure(1, weight=1)

        font_size = 14 if self.root.winfo_screenwidth() >= 1920 else (14 if self.root.winfo_screenwidth() >= 1366 else 12)

        ctk.CTkLabel(color_frame, text="COLOR CODE", font=('Helvetica', font_size, 'bold'),
                     text_color="#333333").grid(row=0, columnspan=2, pady=5)

        ctk.CTkLabel(color_frame, text="Read", font=('Helvetica', 11, 'bold'), fg_color="#90EE90", corner_radius=7, width=70).grid(row=1, column=0, padx=4, pady=4)
        ctk.CTkLabel(color_frame, text="Write", font=('Helvetica', 11, 'bold'), fg_color="#ADD8E6", corner_radius=7, width=70).grid(row=1, column=1, padx=3, pady=4)

    def create_responsive_config_area(self):
        """Create responsive configuration area with scrolling"""
        config_frame = ctk.CTkFrame(self.main_container, border_width=2, border_color="#dddddd", 
                                  corner_radius=8, fg_color="#ffffff")
        config_frame.grid(row=2, column=0, sticky="nsew", padx=5, pady=2)
        
        # Configure config frame grid
        config_frame.grid_rowconfigure(0, weight=1)
        config_frame.grid_columnconfigure(0, weight=1)
        
        # Create a canvas and scrollbars
        self.config_canvas = tk.Canvas(config_frame, bg="white", highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(config_frame, orient="vertical", command=self.config_canvas.yview)
        h_scrollbar = ttk.Scrollbar(config_frame, orient="horizontal", command=self.config_canvas.xview)
        
        # Create scrollable frame
        self.scrollable_frame = ctk.CTkFrame(self.config_canvas, fg_color="white")
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.config_canvas.configure(scrollregion=self.config_canvas.bbox("all")))
        
        # Configure scrollable frame for responsive content
        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        
        # Configure canvas
        self.config_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.config_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout for scrollable area
        self.config_canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Mouse wheel bindings
        self.config_canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.config_canvas.bind_all("<Shift-MouseWheel>", self._on_shift_mousewheel)
        
        # Bind canvas resize to update scroll region
        self.config_canvas.bind('<Configure>', self._on_canvas_configure)
        
        # Load configuration fields
        self.entry_lists1 = self.commonRead(self.scrollable_frame, 'full_config.json', "Full Configuration")

        # After all values are set in AlgoToSetValue, update originals to match GUI
        self.entry_original_values = []
        for entry_row in self.entry_lists1:
            originals = []
            for entry in entry_row:
                if isinstance(entry, str):
                    originals.append(entry)
                else:
                    originals.append(entry.get())
                self.entry_original_values.append(originals)

    # reading the file inside the program
    def get_json_file_path(self, filename):
        return os.path.join(getattr(sys, '_MEIPASS', script_dir), filename)

    def create_responsive_action_buttons(self):
        """Create responsive action buttons"""
        action_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        action_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=(5, 10))
        
        # Configure button grid for responsive layout
        for i in range(6):  # 6 buttons
            action_frame.grid_columnconfigure(i, weight=1, uniform="button")
        
        # Responsive button sizing
        screen_width = self.root.winfo_screenwidth()
        if screen_width >= 3840:  # 4K
            btn_font_size = 16
            btn_height = 50
        elif screen_width >= 1920:  # Full HD
            btn_font_size = 14
            btn_height = 44
        elif screen_width >= 1366:  # 1366x768 - compact but readable
            btn_font_size = 13
            btn_height = 40
        else:  # Smaller screens
            btn_font_size = 12
            btn_height = 38
        
        btn_style = {
            'font': ('Helvetica', btn_font_size, 'bold'), 'height': btn_height, 'corner_radius': 6, 'fg_color': "royalblue", 'hover_color': "indigo", 'text_color': "white"}
        
        # Create buttons with responsive grid layout
        buttons = [
            ("Write", self.processButtonSend),
            ("Read", self.processButtonReceive),
            ("Upload", self.open_file),
            ("Export", self.write_excelbutton),
            ("Zero Angle Estimate", self.start_continuous),
            ("Screenshot", self.capture_screenshot)
        ]
        
        for i, (text, command) in enumerate(buttons):
            btn = ctk.CTkButton(action_frame, text=text, command=command, **btn_style)
            btn.grid(row=0, column=i, padx=5, sticky="ew")

            # Assign the 'Zero Angle Estimate' button to self.button_zero
            if text == "Zero Angle Estimate":
                self.button_zero = btn

    #star continuous
    def start_continuous(self):
        if self.button_zero is None: 
            print("Error: 'Zero Angle Estimate' button not found.")
            return
        
        def task():
            self.button_zero.configure(state='disabled', fg_color="#6c757d")
            self.dialog.configure(text="Zero Angle in progress...", text_color="blue")
            try:
                self.receiveRandom()  # Assuming this is a blocking call
            except Exception as e:
                self.dialog.configure(text=f"Zero Angle Estimate Error: {e}", text_color="red")
            finally:
                self.button_zero.configure(state='normal', fg_color="#2a5885")
                self.dialog.configure(text="Zero Angle finished.", text_color="green") 

        thread = threading.Thread(target=task)
        thread.start()

    #capture screenshot
    def capture_screenshot(self):
        # Ask user where to save with default filename
        default_filename = f"tuner_screenshot_{time.strftime('%Y_%m_%d_%H-%M-%S')}.png"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".png", 
            filetypes=[("PNG files", "*.png")], 
            title="Save Screenshot As",
            initialfile=default_filename)
        
        if file_path:
            # Get window coordinates
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            width = self.root.winfo_width()
            height = self.root.winfo_height()
                
            # Capture and save
            screenshot = pyautogui.screenshot(region=(x, y, width, height))
            screenshot.save(file_path)

    def show_help(self):
        self.help_window = ctk.CTkToplevel(self.root)
        self.help_window.title("Help")
        self.help_window.geometry("500x480")
        self.help_window.resizable(False, False)
        self.help_window.grab_set()
        self.help_window.focus()
        self.help_window.transient(self.root)
    
        help_text = """
        🛠️ MCU Tuner V6.0 – Quick Help Guide
    
        1. COM Port: Select the correct serial port for communication.
        2. PIC ID: Read the unique ID of the connected MCU.
        3. Firmware: Displays the current firmware version.
        4. Status: Shows operation results (e.g., Read Successful, Connection Error).
        5. Color Code: Visual indicator(Read Successful, Write Successful).
        
        🔹  Buttons & Functions
        
        > Read : Reads data from the MCU (PIC ID + Firmware).

        > Write : Writes current settings to the MCU.

        > Upload : Loads configuration from an Excel file.

        > Export : Saves current data to an Excel file.

        > Zero Angle Estimate: Calculates and sets zero angle offset.

        > Screenshot: Takes a screenshot of the current window."""
    
        #For more help, visit our documentation.
        

        text_frame = ctk.CTkFrame(self.help_window, fg_color="white")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
        help_label = ctk.CTkLabel(text_frame, text=help_text,
                                  font=('Helvetica', 13), justify='left', anchor='w', wraplength=480)
        help_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
        #doc_link = ctk.CTkLabel(text_frame, text="Open Documentation", text_color="blue",
          #                      cursor="hand2", font=('Helvetica', 12, 'underline'))
        #doc_link.pack(pady=(0, 10))
        #doc_link.bind("<Button-1>", lambda e: webbrowser.open("https://docs.example.com"))

    def detect_serial_ports(self):
        self.ports = serial.tools.list_ports.comports()
        if self.ports:
            return [port.device for port in serial.tools.list_ports.comports()]
        else:
            return ['No ports available']

    # reading the xlsx side to entry list
    def open_file(self):
        filename = filedialog.askopenfilename(title="Open a File", filetype=(("xlxs files", ".*xlsx"),
                                                                             ("All Files", "*.")))
        if filename:
            try:
                filename = r"{}".format(filename)
                df = openpyxl.load_workbook(filename)
                sheet = df.active
                max_row = sheet.max_row - 2
                rE = 0
                le = 0
            
                # Iterate over rows up to 10
                for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=8):
                    row_data = []
                
                    # Iterate over cells in the row
                    for cell in row: 
                        if cell.column_letter == "B" or cell.column_letter == "D" or cell.column_letter == "F" or cell.column_letter == "H":
                            row_data.append(cell.value)

                    # reading the data and values
                    if le < len(self.entry_lists1):
                        self.AlgoToSetValue(self.entry_lists1[le], row_data)
                    else:
                        print(f"Warning: entry_lists1 index {le} is out of range.")
                    rE += 1
                    le += 1

                    # If the file is successfully processed, show "Upload successful"
                self.dialog.configure(text="Upload successful", font=('Helvetica', 17, 'bold'), fg_color="#ADD8E6")

            except ValueError:
                self.dialog.configure(text="File could not be opened")
        
            except FileNotFoundError:
                self.dialog.configure(text="File Not Found")

    # showing the random data that has to be received from the code
    def receiveRandom(self):
        rs = None
        self.is_running = True
        start_time = time.time()
    
        if self.uartState:
            while self.is_running:
                rs = self.ser.read(20)
                if time.time() - start_time > 12:
                    break
                if rs:
                    self.randomVari(rs, self.entry_lists1[2])
                    break

            if self.is_running:
                self.is_running = False
                self.button3['text'] = "Zero Angle Estimate"
            else:
                self.button3['text'] = "Zero Angle Estimate"


    # reading the picId
    def readFrame(self):
        if self.uartState:
                
                self.ser.write(bytes.fromhex('20012F'))
                bval = self.ser.read(10)
                
                if len(bval) < 10:
                    self.dialog.configure(text="Incomplete PIC ID response", font=('Helvetica', 14, 'bold'))
                    return

                # Check if the expected response length is reached
                if len(bval) >= 10:
                    # Validate the first and last bytes
                    first_byte = bval[0]
                    last_byte = bval[-1]

                    if bval.hex().startswith("20") and len(bval) == 10:
                        if first_byte == 0x20 and last_byte == 0x2f:
                            self.flag = 1
                        else:
                            print(f"Received PIC ID response is Invalid: {bval.hex()}")

                unpacked_data = bval[1:-1]
                h1 = (unpacked_data[:4]).hex()
                h1 = h1[::-1]
                st1 = h1[1] + h1[0] + h1[3] + h1[2] + h1[5] + h1[4] + h1[7] + h1[6]
                h2 = (unpacked_data[4:]).hex()
                h2 = h2[::-1]
                st2 = h2[1] + h2[0] + h2[3] + h2[2] + h2[5] + h2[4] + h2[7] + h2[6]

                formatted_pic_id = st2.zfill(4) + " " + st1.zfill(4)

                # update (self.label2) with h1 value
                self.label2.configure(state='normal')
                self.label2.delete(0, tk.END)
                self.label2.insert(0, st1)
                self.label2.configure(font=('GoudyOLStBT', 14, 'bold'), justify='center', state='readonly')

                # update (self.label3) with h2 value
                self.label3.configure(state='normal')
                self.label3.delete(0, tk.END)
                self.label3.insert(0, st2)
                self.label3.configure(font=('GoudyOLStBT', 14, 'bold'), justify='center', state='readonly')

                # Small delay before sending the next command
                time.sleep(0.1)

                # Read firmware version
                self.ser.timeout = 1
                self.ser.write(bytes.fromhex('30013F'))
                bval_fw = self.ser.read(10)
                
                if len(bval_fw) < 10:
                    self.dialog.configure(text="Incomplete Firmware", font=('Helvetica', 14, 'bold'))
                    return

                # Check if the expected response length is reached
                if len(bval_fw) >= 10:
                    # Validate the first and last bytes
                    first_byte = bval_fw[0]
                    last_byte = bval_fw[-1]

                    if bval_fw.hex().startswith("30") and len(bval_fw) == 10:
                        if first_byte == 0x30 and last_byte == 0x3f:
                            self.flag = 1
                        else:
                            print(f"Received Firmware ID response is Invalid: {bval_fw.hex()}")

                unpacked_data_fw = bval_fw[1:-1]
                #print(f"Unpacked firmware version data: {unpacked_data_fw.hex()}")
                fw1 = (unpacked_data_fw[:4]).hex()
                fw1 = fw1[::-1]
                st_fw1 = fw1[1] + fw1[0] + fw1[3] + fw1[2] + fw1[5] + fw1[4] + fw1[7] + fw1[6]
                fw2 = (unpacked_data_fw[4:]).hex()
                fw2 = fw2[::-1]
                st_fw2 = fw2[1] + fw2[0] + fw2[3] + fw2[2] + fw2[5] + fw2[4] + fw2[7] + fw2[6]

                formatted_fw_version = st_fw2.zfill(4) + " " + st_fw1.zfill(4)

                # update firmware labels
                self.firmware2.configure(state='normal')
                self.firmware2.delete(0, tk.END)
                self.firmware2.insert(0, st_fw1)
                self.firmware2.configure(font=('GoudyOLStBT', 14, 'bold'), justify='center', state='readonly')

                self.firmware3.configure(state='normal')
                self.firmware3.delete(0, tk.END)
                self.firmware3.insert(0, st_fw2)
                self.firmware3.configure(font=('GoudyOLStBT', 14, 'bold'), justify='center', state='readonly')                
        else:
            self.dialog.configure(text="Not in connect", font=('Helvetica', 14, 'bold'), text_color='red')

        # for connecting the process button with port
    def processButtonSS(self):
        if not self.uartState:
            self.ser.port = self.selected_port.get()
            self.ser.baudrate = 9600
            self.ser.timeout = 1
            try:
                self.ser.open()
            except:
                self.dialog.configure(
                    text="Can't open the Port", font=('Helvetica', 17, 'bold'), text_color="red")
        
            if self.ser.is_open:
                self.buttonSS.configure(fg_color="lightsteelblue")  # Set clicked color
                self.buttonSS.update_idletasks()
                self.dump()
                print("Dump command sent")
            
                if self.open == 1:  # open success
                    self.uartState = True

                    if self.flag2 == 0:
                        self.dialog.configure(text="Busy", font=('Helvetica', 17, 'bold'), text_color="blue")
                
                    #reading from the serial port
                    self.serial_operations()

                    #disabled the start button
                    self.buttonSS.update_idletasks()
                    self.disable_button(self.buttonSS)
                    self.enable_button(self.buttonStop)

    def stopButtonSS(self):
        if self.uartState:
            self.ser.close()
            self.buttonSS["text"] = "START"
            self.uartState = False
            self.dialog.configure(text="Port Closed", font=('Helvetica', 17, 'bold'), fg_color='red', text_color='white')

            #disabled the STOP button after its click
            self.disable_button(self.buttonStop)
            self.enable_button(self.buttonSS)

    def serial_operations(self):
        self.processReceived()
        self.readFrame()
    
        self.flag2 = 1
        if self.flag2 == 1:
            self.dialog.configure(
                text="Read Successful", font=('Helvetica', 17, 'bold'), text_color='black')
            self.flag2 = 0
            self.open = 0

    # Assuming you have these functions defined elsewhere
    def on_enter(self, button, color):
        button.configure(fg_color=color)

    def on_leave(self, button, color):
        button.configure(fg_color=color)

    # Disable button function 
    def disable_button(self, button):
        button.configure(state="disabled", fg_color="grey", text_color="darkgrey")
        button.unbind("<Enter>")  # disable hover effect
        button.unbind("<Leave>")

    # Enable button function 
    def enable_button(self, button):
        if button == self.buttonSS:
            button.configure(state="normal", fg_color="royalblue", text_color="white")
            self.buttonSS.bind("<Enter>", lambda event, button=self.buttonSS: self.on_enter(button, "blue"))
            self.buttonSS.bind("<Leave>", lambda event, button=self.buttonSS: self.on_leave(button, "royalblue"))
        elif button == self.buttonStop:
            button.configure(state="normal", fg_color="red", text_color="white")
            self.buttonStop.bind("<Enter>", lambda event, button=self.buttonStop: self.on_enter(button, 'red2'))
            self.buttonStop.bind("<Leave>", lambda event, button=self.buttonStop: self.on_leave(button, 'red'))

    def on_port_selected(self, *args):
        selected_port = self.selected_port.get()
        if selected_port != "None":  
            self.buttonSS.configure(state="normal", text_color="white")
            self.dialog.configure(text="  Serial Port Connected  ", font=('Helvetica', 17, 'bold'), fg_color="#ADD8E6", text_color='black')
        else:  
            self.buttonSS.configure(fg_color="grey", text_color="white")

    def dump(self):
        start_time = time.time()
        self.ser.write(bytes.fromhex('10011f'))
    
        response = None
        while True:
            if self.ser.in_waiting:
                response = self.ser.read(3)
                break
            if time.time() - start_time > self.ser.timeout:  # Timeout of 5 seconds
                # print("no response")
                break

        # process the response
        if response:
            self.open = 1
        else:
            self.dialog.configure(text="Connection Error", font=('Helvetica', 17, 'bold'), fg_color='red', text_color='white')

    # stopping the function
    def stop_receiveRandom(self):
        self.is_running = False

    # for receiving the values from the uart
    def processButtonReceive(self):
        
        if self.uartState:
            if self.flag2 == 0:
                self.dialog.configure(text="Busy", font=('Helvetica', 17, 'bold'), fg_color='orange')

            self.is_running = False
            self.dialog.configure(text="Reading...", font=('Helvetica', 17, 'bold'), text_color='white', fg_color='indigo')

            value = []
            for data in self.entry_lists1:
                for entry in data:
                    if isinstance(entry, str):
                        value.append(entry[2:]+"00")

            self.received_data(value, self.entry_lists1)

            if self.flag2 == 1:
                self.dialog.configure(text="  Read Successful  ", font=('Helvetica', 17, 'bold'), fg_color="#90EE90", text_color='black')
                self.flag2 = 0
            else:
                self.dialog.configure(text="  Connection Error  ", font=('Helvetica', 17, 'bold'), fg_color='red', text_color='white')

        else:
            self.dialog.configure(text="  Read-> Not in Connect  ", font=('Helvetica', 17, 'bold'))

    # receving the data if the button is pressed
    def processReceived(self):

        if (self.uartState):
            self.dialog.configure(text="  Busy  ", font=('Helvetica', 17, 'bold'))
            value = []
            for data in self.entry_lists1:
                for entry in data:
                    if isinstance(entry, str):
                        value.append(entry[2:]+"00")

            self.received_data(value, self.entry_lists1)

            if(self.flag2 == 1):
                self.dialog.configure(text="  Read Successful  ", font=('Helvetica', 17, 'bold'), fg_color="#90EE90", text_color='black')
                self.flag2 = 0
        else:
            self.dialog.configure(text="  Received-> Not In Connect  ")

    # sending the data to the uart
    def processButtonSend(self):
            
        if (self.uartState):
            self.dialog.configure(text="  Busy  ", font=('Helvetica', 17, 'bold'))

            self.is_running = False
            self.readEntry(self.entry_lists1)

            if self.flag == 0:
                self.dialog.configure(text="  Connection Error  ",  font=('Helvetica', 17, 'bold'), fg_color="red", text_color='white')
            else:
                self.dialog.configure(text="  Write Successful  ", font=('Helvetica', 17, 'bold'), fg_color="#ADD8E6", text_color='black')
                self.flag = 0
        else:
            self.dialog.configure(text="  Sent-> Not in Connect  ",  font=('Helvetica', 17, 'bold'))

    def readEntry(self, data):
        # get the values from the entry widgets
        values = []
        for entry_list in data:
            frame_values = []
            for entry in entry_list:
                if isinstance(entry, str):
                    frame_values.append(entry)
                else:
                    frame_values.append(float(entry.get()))
            values.append(frame_values)
            

        # converting the string and float values to hexvalues and sent to bytes
        value = b''
        for i in values:
            value = b''
            bv = b''
            for j in i:
                # print(j)
                if isinstance(j, str):
                    bv = bytes.fromhex(j[2:])
                    value = value + bytes.fromhex(j[2:])
                else:
                    value = value + struct.pack('<f', j)
            # bytes values
            value = value + bv
            
            self.ser.write(bytes(value))
            start_time = time.time()
            res = None
            # sending and checking whether the values has been sent or not if
            # is then sent the next
            while True:
                if self.ser.in_waiting > 0:
                    res = self.ser.read(20)
                    #
                    # if bytes(value) == res:
                    #     print('True')
                    # else:
                    #     print('False')
                    break
                if time.time() - start_time > self.ser.timeout:
                    break
            if res:
                self.open = 1
                # break
            else:
                break
        # entries that has been sent
        if self.open == 1:
            for row_idx, entry_row in enumerate(data):
                for col_idx, entry in enumerate(entry_row):
                    if isinstance(entry, str):
                        continue
                    else:
                        try:
                            original = self.entry_original_values[row_idx][col_idx]
                        except IndexError:
                            original = None
                        if entry.get() != original:
                            entry.configure(fg_color="#2fadd6")  # Highlight edited
                            self.entry_original_values[row_idx][col_idx] = entry.get()
                        else:
                            entry.configure(fg_color="#ADD8E6")    # Reset unchanged to white
            self.flag = 1
            self.open = 0
        else:
            self.flag = 0
            self.open = 0

    def _on_canvas_configure(self, event):
        """Handle canvas resize to update scroll region"""
        self.config_canvas.configure(scrollregion=self.config_canvas.bbox("all"))
        
        # Update the scrollable frame width to match canvas width
        canvas_width = event.width
        self.config_canvas.itemconfig(self.config_canvas.find_all()[0], width=canvas_width)

    def _on_mousewheel(self, event):
        """Handle vertical mouse wheel scrolling"""
        self.config_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_shift_mousewheel(self, event):
        """Handle horizontal mouse wheel scrolling"""
        self.config_canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")

    # Updated commonRead method with responsive design
    def commonRead(self, parent_frame, data, title):
        filename = self.get_json_file_path(data)
        with open(filename, 'r') as f:
            data = json.load(f)
            value = data['instructions']
            
            # Create a collapsible frame for this configuration section
            collapsible_frame = CollapsibleFrame(parent_frame, text=title)
            collapsible_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=0)
            
            entry_lists = []
            
            # Create a grid layout for the fields
            grid_frame = ctk.CTkFrame(collapsible_frame.sub_frame, fg_color="white")
            grid_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=0)

            # Configure grid columns (8 columns total) - Make them responsive
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            if screen_width >= 1366:
                min_col_size = 110
            else:
                min_col_size = 90
            
            # Configure grid columns (8 columns total) - Make them responsive
            for i in range(8):
                grid_frame.grid_columnconfigure(i, weight=1, uniform="col", minsize=min_col_size)

            num_rows_to_display = len(value) # Total number of rows you will create

            # Dynamic row sizing parameters based on screen height
            if screen_height >= 1080: # Full HD and above
                dynamic_font_size = 13
                dynamic_ipady = 6
                dynamic_pady = 2
                dynamic_row_minsize = 45 
            elif screen_height >= 1050: 
                dynamic_font_size = 11
                dynamic_ipady = 4
                dynamic_pady = 2
                dynamic_row_minsize = 40
            elif screen_height >= 900: 
                dynamic_font_size = 11
                dynamic_ipady = 4
                dynamic_pady = 2
                dynamic_row_minsize = 45
            elif screen_height >= 960: 
                dynamic_font_size = 11
                dynamic_ipady = 10
                dynamic_pady = 3
                dynamic_row_minsize = 60
            elif screen_height >= 800: 
                dynamic_font_size = 11
                dynamic_ipady = 2
                dynamic_pady = 2
                dynamic_row_minsize = 30
            elif screen_height >= 768: # 1366x768 or similar
                dynamic_font_size = 11
                dynamic_ipady = 2 # Reduced internal padding
                dynamic_pady = 1  # Reduced external padding
                dynamic_row_minsize = 30 # Reduced row minsize
            elif screen_height >= 720: 
                dynamic_font_size = 11
                dynamic_ipady = 1
                dynamic_pady = 1
                dynamic_row_minsize = 28
            else: # Smaller screens (e.g., netbooks, very small resolutions)
                dynamic_font_size = 11
                dynamic_ipady = 1 # Even less internal padding
                dynamic_pady = 0  # No external padding between rows
                dynamic_row_minsize = 28 # Even smaller row minsize
            
            row = 0
            for i in range(num_rows_to_display):
                val = []
                ind = value[i]['fields']
                val.append(value[i]['id'])
                
                # Make the row expandable
                grid_frame.grid_rowconfigure(row, weight=1, minsize=dynamic_row_minsize)
                
                for j in range(len(ind)):
                    
                    label = ctk.CTkLabel(grid_frame, text=ind[j]['field1_name'], font=('Helvetica', dynamic_font_size, 'bold'), fg_color='darkslategrey',
                                         text_color='white', anchor='w', corner_radius=7, wraplength=180)
                    label.grid(row=row, column=j*2, sticky='nsew', padx=(5, 1), pady=dynamic_pady, ipady=dynamic_ipady)
                    
                    # Entry field - Make it responsive
                    var = tk.StringVar()
                    entry = ctk.CTkEntry(grid_frame, textvariable=var, font=('Helvetica', dynamic_font_size, 'bold'), border_color="black", corner_radius=7)
                    entry.custom_name = ind[j]["field1_name"]
                    entry.grid(row=row, column=j*2+1,  sticky='nsew', padx=(1, 5), pady=dynamic_pady, ipady=dynamic_ipady)
                    
                    val.append(entry)
                
                entry_lists.append(val)
                row += 1
                
            return entry_lists

    # # Receiving the values
    def received_data(self, vbyte, entry_list):
        le = 0
        flag = 0
        
        for i in vbyte:
            start_time = time.time()
            byte_to_send = bytes.fromhex(i)
            check = True
            
            while check:
                self.ser.write(byte_to_send)
                
                while True:
                    if self.ser.in_waiting:
                        res = self.ser.read(20)
                        print(res)

                        # Check if the expected response length is reached
                        if len(res) >= 20:
                            first_byte = res[0]      # Validate the first and last 
                            last_byte = res[-1]

                            if res.hex().startswith("ff") and len(res) == 20:
                                if first_byte == 0xff and last_byte != 0x00:
                                    self.flag = 1
                                else:
                                    print(f"Received Invalid response: {res()}")
                        
                        val = binascii.hexlify(struct.unpack("<2s", res[0:2])[0]).decode() + "00"
                        val = val.upper()
                        
                        if val == i:
                            check = False
                        break
                    
                    if time.time() - start_time > self.ser.timeout:
                        flag = 1
                        check = False
                        break
                    
            if flag == 1:
                self.dialog.configure(text="Connection Error", font=('Helvetica', 17, 'bold'), fg_color='red', text_color='white')
                self.flag2 = 0
                break
            
            else:
                actual = self.AlgoToRead(res)
                self.AlgoToSetValue(entry_list[le], actual)
                # After all values are set in AlgoToSetValue, update originals to match GUI
                self.entry_original_values = []
                for entry_row in entry_list:
                    originals = []
                    for entry in entry_row:
                        if isinstance(entry, str):
                            originals.append(entry)
                        else:
                            originals.append(entry.get())
                    self.entry_original_values.append(originals)
                le += 1
                self.open = 0
                self.flag2 = 1

    # packing the values for the floats to hex then bytes
    def AlgoToRead(self, data):
        actual = ["{:.4f}".format(struct.unpack('<f', data[i:i + 4])[0])
                  for i in range(2, len(data) - 4, 4)]
        print(actual)
        return actual

    # setting the value according to the entries that are givens
    def AlgoToSetValue(self, entry_list, data):
        
        for i in range(1, len(entry_list)):
            #print("Entry list name ===> ",entry_list[i].custom_name)
    
            if entry_list[i].custom_name in ["kmph/rpm", "rpm kp", "rpm ki"]:
                formatted_value = "{:.4f}".format(float(data.pop(0)))
            else:
                formatted_value = "{:.2f}".format(float(data.pop(0)))
            entry_list[i].delete(0, tk.END)
            entry_list[i].insert(0, formatted_value)
            entry_list[i].configure(fg_color="#90EE90", text_color="black")

    def randomVari(self, rs, entry_list):
        actual = self.AlgoToRead(rs)
        for i in range(1, len(entry_list)):
            entry_list[i].delete(0, tk.END)
            entry_list[i].insert(0, actual.pop(0))
            entry_list[i].configure(fg_color="#ADD8E6", text_color="black")

    def common_write(self, data, list, col, worksheet, entry_list):
        filename = self.get_json_file_path(data)
        with open(filename, 'r') as f:
            data = json.load(f)
            value = data['instructions']
            le = 0
            for i in range(len(value)):
                ind = value[i]['fields']
                row = 0
                # parameters names
                for j in range(len(ind)):
                    worksheet.write(list[row] + str(col), ind[j]['field1_name'])
                    row += 2
                row = 1
                # values
                for j in entry_list[le]:
                    if isinstance(j, str):
                        continue
                    else:
                        worksheet.write(list[row] + str(col), float(j.get()))
                        row += 2
                le += 1
                col += 1
        return col, worksheet

    def write_excelbutton(self):
        now = datetime.now()

        # Default file name with timestamp
        default_filename = 'MCU_Parameters' + now.strftime("D%d_%m_%Y_T%H-%M") + '.xlsx'

        # Open a file save dialog
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx")],
                                             title="Save the Excel file", initialfile=default_filename)

        # If a file path is selected, create and save the Excel file
        if save_path:
            workbook = xlsxwriter.Workbook(save_path)
            worksheet = workbook.add_worksheet()

        # define the list of column letters
        column_letters = ["A", "B", "C", "D", "E", "F", "G", "H"]

        # call the common_write method to populate the worksheet with data
        col, worksheet = self.common_write('full_config.json', column_letters, 1, worksheet, self.entry_lists1)

        # determine the last row index
        last_row_index = col + 1  # +1 to account for 0-based indexing

        # write Date and PicID at the end of the rows
        worksheet.write(f"A{last_row_index + 1}", "Date")
        worksheet.write(f"B{last_row_index + 1}", now.strftime("%d/%m/%Y %H:%M:%S"))
        worksheet.write(f"A{last_row_index + 2}", "PicID")
        worksheet.write(f'B{last_row_index + 2}', self.label2.get())
        worksheet.write(f'C{last_row_index + 2}', self.label3.get())
        worksheet.write(f'A{last_row_index + 3}', "Firmware Version")
        worksheet.write(f'B{last_row_index + 3}', self.firmware2.get())
        worksheet.write(f'C{last_row_index + 3}', self.firmware3.get())

        workbook.close()

        # Update the dialogue box with the message "Values printed"
        self.dialog.configure(text="Values printed", font=('Helvetica', 17, 'bold'), fg_color="#ADD8E6", text_color='black')


# Updated CollapsibleFrame class
class CollapsibleFrame(ctk.CTkFrame):
    def __init__(self, parent, text="", *args, **kwargs):
        ctk.CTkFrame.__init__(self, parent, *args, **kwargs)
        self.configure(fg_color="white")
        
        # Configure the main frame to be expandable
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.show = tk.BooleanVar(value=True)
        
        # Title frame - Fixed height but expandable width
        self.title_frame = ctk.CTkFrame(self, fg_color="#f8f9fa", corner_radius=4)
        self.title_frame.grid(row=0, column=0, sticky='ew', padx=0, pady=0)
        self.title_frame.grid_columnconfigure(0, weight=1)

        # Responsive font sizing for title_label (already in your code)
        screen_width = parent.winfo_screenwidth() # Use parent's screen width for consistency
        if screen_width >= 1920:
            title_font_size = 13
        elif screen_width >= 1366:
            title_font_size = 12
        else:
            title_font_size = 11
        
        # Title label - Responsive text sizing
        self.title_label = ctk.CTkLabel(self.title_frame, text=text, font=('Helvetica', title_font_size, 'bold'), text_color="#2a5885", anchor='w')
        self.title_label.grid(row=0, column=0, sticky='w', padx=10)
        
        # Toggle button - Fixed size but positioned responsively
        self.toggle_button = ctk.CTkButton(self.title_frame, width=30, height=30, text="-", font=('Helvetica', 15), fg_color="transparent", text_color="black", hover=False, command=self.toggle)
        self.toggle_button.grid(row=0, column=1, sticky='e', padx=10)
        
        # Sub frame - Expandable content area
        self.sub_frame = ctk.CTkFrame(self, fg_color="white")
        self.sub_frame.grid_rowconfigure(0, weight=1)
        self.sub_frame.grid_columnconfigure(0, weight=1)
        
        # Show sub_frame on startup if show is True
        if self.show.get():
            self.sub_frame.grid(row=1, column=0, sticky='nsew', pady=(5, 10))
        
    def toggle(self):
        if self.show.get():
            self.sub_frame.grid_forget()
            self.toggle_button.configure(text="+")
            self.show.set(False)
        else:
            self.sub_frame.grid(row=1, column=0, sticky='nsew', pady=(5, 10))
            self.toggle_button.configure(text="-")
            self.show.set(True)

# main for running the project
if __name__ == "__main__":
    root = ctk.CTk()
    root.update()
    root.state('zoomed')
    app = MainGui(root)
    app.ser.close()
    root.mainloop()            
